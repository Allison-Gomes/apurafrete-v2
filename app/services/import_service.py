'''
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📁 ARQUIVO : import_service.py
📦 MÓDULO  : Embarque / Importação de NF
🎯 OBJETIVO: Service que orquestra a importação de NFs:
              0. Valida existência do embarque
              1. Valida linhas via validacao_service
              2. Deduplica vs. banco (Decisão #44)
              3. Persiste via repository (Decisão #45)
              4. Commit/rollback centralizado
             + Funções auxiliares:
              - parse_planilha_nf(): .xlsx → list[NFImportRow]
              - carregar_catalogo_produtos(): DB → dict[SKU, ProdutoInfo]
📐 REGRA    : - Decisão #44: dedup no service
               (consulta repository → filtra → insere só novas)
             - Decisão #45: commit direto no service
               (try/commit/except/rollback, sem UnitOfWork)
             - Decisão #33: duplicatas ignoradas silenciosamente
             - Decisão #41: repository só flush()
             - Decisão #39: peso derivado do catálogo
               (já resolvido no validacao_service)
🔗 DEPENDE  : app/services/validacao_service.py
             app/repositories/nota_fiscal_repository.py
             app/repositories/embarque_repository.py
             app/schemas/nota_fiscal.py
             app/models/produto.py
             app/exceptions/embarque_exceptions.py
📅 CRIADO   : 11/07/2026
📅 ATUALIZ. : 28/07/2026 — Decisão #75. Alinhamento do
                           EXCEL_COLUMN_MAP aos nomes reais dos
                           campos de NFImportRow:
                             documento       → numero_nf
                             centro_de_custo → centro_custo
                           Import corrigido para app.schemas.nota_fiscal.
                           +normalização de células (strip / None).
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
'''

from __future__ import annotations

from io import BytesIO
from uuid import UUID

from openpyxl import load_workbook
from pydantic import BaseModel, Field, ValidationError
from sqlalchemy.orm import Session

from app.exceptions.embarque_exceptions import EmbarqueNaoEncontradoError
from app.models.produto import Produto
from app.repositories.embarque_repository import EmbarqueRepository
from app.repositories.nota_fiscal_repository import NotaFiscalRepository
from app.schemas.nota_fiscal import (
    NFImportError,
    NFImportRow,
    NotaFiscalCreate,
)
from app.services.validacao_service import (
    ProdutoInfo,
    processar_importacao,
)


# ─────────────────────────────────────────────────
# 🗺️ MAPEAMENTO: Cabeçalho Excel → Campo Schema
# ─────────────────────────────────────────────────
# ⚠️ As chaves são os cabeçalhos LITERAIS da planilha (Seção 4.1).
#    Os valores são os nomes EXATOS dos campos de NFImportRow.
#    Qualquer divergência aqui derruba a importação inteira na
#    linha 2, pois o Pydantic descarta a chave desconhecida e
#    acusa o campo obrigatório como ausente.
EXCEL_COLUMN_MAP: dict[str, str] = {
    "DOCUMENTO":              "numero_nf",
    "COD CLIENTE":            "cod_cliente",
    "CLIENTE DESTINO":        "cliente_destino",
    "CNPJ DESTINO":           "cnpj_destino",
    "CIDADE - UF DESTINO":    "cidade_uf_destino",
    "COD REMETENTE":          "cod_remetente",
    "CLIENTE REMETENTE":      "cliente_remetente",
    "CNPJ REMETENTE":         "cnpj_remetente",
    "CIDADE - UF REMETENTE":  "cidade_uf_remetente",
    "COD PRODUTO":            "cod_produto",
    "QTD CX":                 "qtd_cx",
    "OBSERVAÇÃO":             "observacao",
    "CENTRO DE CUSTO":        "centro_custo",
}

# Colunas obrigatórias — se faltar alguma, a planilha é rejeitada.
# Nomes de CAMPO (pós-mapeamento), não de cabeçalho.
REQUIRED_COLUMNS: list[str] = [
    "numero_nf", "cod_cliente", "cliente_destino", "cnpj_destino",
    "cidade_uf_destino", "cod_remetente", "cliente_remetente",
    "cnpj_remetente", "cidade_uf_remetente", "cod_produto", "qtd_cx",
]

# Rótulo legível por campo, para mensagens de erro ao usuário.
_FIELD_TO_HEADER: dict[str, str] = {v: k for k, v in EXCEL_COLUMN_MAP.items()}


# ─────────────────────────────────────────────────
# 📊 SCHEMA LOCAL: ImportLoteResult
# Consolidado final: validação + dedup + persistência.
# ─────────────────────────────────────────────────
class ImportLoteResult(BaseModel):
    '''
    🎯 O QUE FAZ:
        Resultado consolidado do lote de importação,
        unificando os dados de validação e deduplicação.

    📐 REGRA DE NEGÓCIO:
        - total_importadas = efetivamente persistidas
          (passaram na validação E não eram duplicatas).
        - total_erros = erros de validação apenas.
        - duplicatas_ignoradas = passaram na validação
          mas já existiam no embarque (não são erro).
        - Σ = total_importadas + total_erros + duplicatas_ignoradas
          = total_linhas da planilha.
    '''
    total_linhas: int = Field(..., ge=0)
    total_importadas: int = Field(..., ge=0)
    total_erros: int = Field(..., ge=0)
    duplicatas_ignoradas: int = Field(default=0, ge=0)
    erros: list[NFImportError] = Field(default_factory=list)


# ══════════════════════════════════════════════════
# 🔧 FUNÇÕES AUXILIARES
# ══════════════════════════════════════════════════

def _normalizar_celula(valor: object) -> object | None:
    '''
    🎯 O QUE FAZ:
        Normaliza o valor bruto de uma célula do Excel antes
        de alimentar o schema.

    📐 REGRA:
        - None permanece None.
        - str é feito strip(); string vazia vira None (para que
          campos opcionais não recebam "" e campos obrigatórios
          falhem corretamente como ausentes).
        - Demais tipos (int, float, datetime, Decimal) passam
          intactos — a coerção é do Pydantic.
    '''
    if valor is None:
        return None
    if isinstance(valor, str):
        limpo = valor.strip()
        return limpo if limpo else None
    return valor


def _linha_vazia(row: tuple) -> bool:
    '''
    🎯 O QUE FAZ:
        Detecta linhas totalmente vazias, comuns no fim de
        planilhas editadas manualmente (openpyxl as devolve
        como tuplas de None).

    📐 REGRA:
        Linha vazia é ignorada silenciosamente, não conta como
        erro nem entra em total_linhas.
    '''
    return all(_normalizar_celula(c) is None for c in row)


def parse_planilha_nf(conteudo: bytes) -> list[NFImportRow]:
    '''
    🎯 O QUE FAZ:
        Lê um arquivo .xlsx (bytes) e converte cada linha
        de dados em um objeto NFImportRow.

    📐 REGRA:
        - A primeira linha da planilha é tratada como cabeçalho.
        - Colunas obrigatórias ausentes → ValueError (router → HTTP 400).
        - Linhas totalmente vazias são ignoradas.
        - Linhas de dados que falham na construção do schema
          → ValueError com número da linha.
        - Colunas opcionais (OBSERVAÇÃO, CENTRO DE CUSTO) são
          aceitas se existirem; caso contrário, None.
        - Não faz validação de negócio — isso é delegado ao
          validacao_service.processar_importacao().

    📥 conteudo (bytes): conteúdo bruto do arquivo .xlsx.

    📤 list[NFImportRow]: linhas parseadas, prontas para validação.

    ⚠️ Levanta ValueError se:
        - Planilha vazia ou sem dados (só cabeçalho).
        - Coluna obrigatória ausente no cabeçalho.
        - Linha de dados inválida (ex.: tipo incompatível).
    '''
    wb = load_workbook(filename=BytesIO(conteudo), read_only=True, data_only=True)

    try:
        ws = wb.active

        if ws is None:
            raise ValueError("Planilha não contém uma aba ativa.")

        rows = list(ws.iter_rows(min_row=1, values_only=True))

        if len(rows) < 2:
            raise ValueError(
                "Planilha vazia ou sem dados. "
                "Esperado: linha 1 = cabeçalho, linhas 2+ = dados."
            )

        # ── Cabeçalho ──────────────────────────────
        raw_headers = [
            str(h).strip().upper() if h is not None else ""
            for h in rows[0]
        ]

        col_index: dict[str, int] = {}
        for idx, header in enumerate(raw_headers):
            field = EXCEL_COLUMN_MAP.get(header)
            if field:
                col_index[field] = idx

        missing = [f for f in REQUIRED_COLUMNS if f not in col_index]
        if missing:
            rotulos = [_FIELD_TO_HEADER.get(f, f) for f in missing]
            raise ValueError(
                f"Colunas obrigatórias ausentes no cabeçalho: "
                f"{', '.join(rotulos)}"
            )

        # ── Linhas de dados ────────────────────────
        import_rows: list[NFImportRow] = []

        for row_num, row in enumerate(rows[1:], start=2):
            if _linha_vazia(row):
                continue

            row_data: dict[str, object | None] = {}

            for field, idx in col_index.items():
                bruto = row[idx] if idx < len(row) else None
                row_data[field] = _normalizar_celula(bruto)

            try:
                import_rows.append(NFImportRow(**row_data))
            except ValidationError as exc:
                raise ValueError(f"Erro na linha {row_num}: {exc}") from exc

        if not import_rows:
            raise ValueError("Planilha não contém nenhuma linha de dados válida.")

        return import_rows

    finally:
        wb.close()


def carregar_catalogo_produtos(db: Session) -> dict[str, ProdutoInfo]:
    '''
    🎯 O QUE FAZ:
        Carrega todos os produtos ativos do banco e monta
        o catálogo SKU → ProdutoInfo, usado pelo import_service
        para derivar peso total na validação.

    📐 REGRA:
        - Apenas produtos ativos (ativo=True) entram no catálogo.
        - Se peso_real_kg for None ou ≤ 0, o produto é ignorado
          (a validação vai rejeitar a linha como SEM_PRODUTO).
        - Chamado pelo router ANTES de invocar importar_nfs(),
          para evitar N+1 queries dentro do loop de validação.

    📥 db (Session): sessão SQLAlchemy ativa.

    📤 dict[str, ProdutoInfo]: chave = SKU, valor = ProdutoInfo.
    '''
    produtos = (
        db.query(Produto)
        .filter(Produto.ativo.is_(True))
        .all()
    )

    catalogo: dict[str, ProdutoInfo] = {}
    for p in produtos:
        if p.peso_real_kg and p.peso_real_kg > 0:
            catalogo[p.sku] = ProdutoInfo(
                sku=p.sku,
                peso_real_kg=p.peso_real_kg,
            )

    return catalogo


# ══════════════════════════════════════════════════
# 🚀 ENTRYPOINT: importar_nfs
# ══════════════════════════════════════════════════

def importar_nfs(
    session: Session,
    linhas: list[NFImportRow],
    embarque_id: UUID,
    catalogo: dict[str, ProdutoInfo],
) -> tuple[list, ImportLoteResult]:
    '''
    🎯 O QUE FAZ:
        Orquestra o fluxo completo de importação de NFs
        em lote, da planilha bruta até a persistência:
          0. Valida existência do embarque
          1. Valida e enriquece as linhas (validacao_service)
          2. Consulta chaves já existentes no embarque (repository)
          3. Filtra duplicatas — ignoradas silenciosamente (#33, #44)
          4. Persiste apenas as NFs novas (repository → flush)
          5. Commit da transação (#45)

    📐 REGRA DE NEGÓCIO:
        - Embarque inexistente/inativo → EmbarqueNaoEncontradoError.
        - Duplicatas NÃO interrompem o lote e NÃO viram erro.
        - Chave de dedup: (numero_nf, serie_nf) dentro do embarque,
          espelhando uq_nf_embarque_numero_serie.
        - Deduplica também DENTRO do próprio lote, evitando violar
          a unique quando a planilha repete a mesma NF.
        - Se zero linhas válidas após validação → retorna vazio
          sem tocar no banco (nem chama o repository).
        - Se todas as válidas forem duplicatas → retorna vazio
          sem insert (só fez a consulta de chaves).
        - Em caso de falha no commit → rollback + relança exceção.
        - O chamador (router) é responsável por abrir/fechar a
          session; este service apenas opera dentro dela.

    📥 PARÂMETROS:
        session (Session): sessão SQLAlchemy síncrona (R1).
        linhas (list[NFImportRow]): linhas brutas da planilha.
        embarque_id (UUID): embarque de destino das NFs.
        catalogo (dict[str, ProdutoInfo]): catálogo SKU → peso
          (já carregado pelo router, evita N+1 no service).

    📤 RETORNO:
        tuple[list, ImportLoteResult]:
          - models persistidos (com id, flushados).
          - resultado consolidado com validação + dedup.

    ⚠️ EXCEÇÕES:
        - EmbarqueNaoEncontradoError: embarque não existe ou inativo.
    '''
    # ── Etapa 0: Validar existência do embarque ──
    embarque_repo = EmbarqueRepository(session)
    if not embarque_repo.existe(embarque_id):
        raise EmbarqueNaoEncontradoError(embarque_id)

    # ── Etapa 1: Validação (pura, sem I/O) ──────
    criadas, resultado_validacao = processar_importacao(
        linhas=linhas,
        embarque_id=embarque_id,
        catalogo=catalogo,
    )

    total_linhas = resultado_validacao.total_linhas
    total_erros = resultado_validacao.total_erros
    erros_validacao = resultado_validacao.erros

    if not criadas:
        # Nenhuma linha passou na validação —
        # nem encosta no banco.
        return [], ImportLoteResult(
            total_linhas=total_linhas,
            total_importadas=0,
            total_erros=total_erros,
            duplicatas_ignoradas=0,
            erros=erros_validacao,
        )

    # ── Etapa 2: Dedup vs. banco + intra-lote (#44) ──
    repo = NotaFiscalRepository(session)

    chaves_existentes = set(repo.buscar_chaves_existentes(embarque_id))

    novas: list[NotaFiscalCreate] = []
    duplicatas = 0
    vistas: set[tuple[str, str | None]] = set()

    for nf in criadas:
        chave = (nf.numero_nf, nf.serie_nf)
        if chave in chaves_existentes or chave in vistas:
            duplicatas += 1  # ignorada silenciosamente (#33)
        else:
            vistas.add(chave)
            novas.append(nf)

    if not novas:
        # Todas as válidas já existiam —
        # sem insert, sem commit.
        return [], ImportLoteResult(
            total_linhas=total_linhas,
            total_importadas=0,
            total_erros=total_erros,
            duplicatas_ignoradas=duplicatas,
            erros=erros_validacao,
        )

    # ── Etapa 3: Persistência (#45) ─────────────
    try:
        objetos = repo.criar_em_lote(novas)   # flush apenas
        session.commit()                       # commit aqui (#45)
    except Exception:
        session.rollback()
        raise

    # ── Etapa 4: Resultado consolidado ──────────
    resultado = ImportLoteResult(
        total_linhas=total_linhas,
        total_importadas=len(novas),
        total_erros=total_erros,
        duplicatas_ignoradas=duplicatas,
        erros=erros_validacao,
    )

    return objetos, resultado
